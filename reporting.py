"""
reporting.py
============
Generate outputs from the causal_analysis pipeline:
  - Excel workbook   three sheets: Double ML, Causal Forest ATE, Backdoor
                     with colour-coded rows (green = significant/good,
                     yellow = caution, orange = problem)
  - CATE tree PDFs   one PDF per (choice_set, transition) for every fitted
                     artifact; saved into cate_trees/significant/ (p < 0.05)
                     or cate_trees/not-significant/ (p ≥ 0.05)
  - PDF report       full narrative report with all tables and plots

Usage
-----
    from reporting import generate_report

    paths = generate_report(
        decisions   = decisions,
        lr_results  = logistic_regression_test(decisions),
        dml_results = double_ml_test(decisions, FEATURES),
        summary     = summary,       # first return value of causal_forest_dml()
        artifacts   = artifacts,     # second return value of causal_forest_dml()
        dr          = conditional_dose_response(artifacts),
        backdoor    = backdoor_check(decisions, FEATURES),
        output_dir  = "output/bpic12",
        file_stem   = "bpic12-dfg",
    )
    # paths == {"excel": "...", "pdf": "...", "trees_dir": "...", "tree_pdfs": [...]}
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import PatternFill, Font, Alignment
from sklearn.tree import plot_tree


# ─────────────────────────────────────────────────────────────────────────────
# Generic helpers
# ─────────────────────────────────────────────────────────────────────────────

def _stringify_lists(df: pd.DataFrame) -> pd.DataFrame:
    """Convert list/tuple cells to comma-separated strings (safe for Excel)."""
    out = df.copy()
    for col in out.columns:
        if out[col].apply(lambda x: isinstance(x, (list, tuple))).any():
            out[col] = out[col].apply(
                lambda x: ", ".join(str(v) for v in x) if isinstance(x, (list, tuple)) else x
            )
    return out


def _render_table(ax, df: pd.DataFrame, title: str = "", fontsize: int = 7):
    """Render a DataFrame as a formatted matplotlib table on ax."""
    ax.axis("off")
    ax.set_title(title, fontsize=9, fontweight="bold", pad=8, loc="left")
    if df.empty:
        ax.text(0.5, 0.5, "(no data)", ha="center", va="center",
                transform=ax.transAxes, color="gray", fontsize=10)
        return
    safe = _stringify_lists(df)
    tbl = ax.table(
        cellText=safe.astype(str).values.tolist(),
        colLabels=list(safe.columns),
        cellLoc="center",
        loc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(fontsize)
    tbl.auto_set_column_width(col=list(range(len(safe.columns))))
    for (r, c), cell in tbl.get_celld().items():
        if r == 0:
            cell.set_facecolor("#2c3e50")
            cell.set_text_props(color="white", fontweight="bold")
        elif r % 2 == 0:
            cell.set_facecolor("#f5f5f5")
        cell.set_edgecolor("#dddddd")
        cell.set_linewidth(0.5)


LEGEND_MAX = 6


def _tree_figsize(tree) -> tuple[float, float]:
    """Return (width, height) inches for a sklearn decision tree figure.

    Width scales with actual leaf count (not 2^depth), capped at 16 inches so
    the PDF stays on a reasonable page.  Height grows with depth.
    """
    n_leaves = tree.get_n_leaves()
    depth    = tree.get_depth()
    w = min(max(8.0, n_leaves * 2.0), 16.0)
    h = max(4.0, depth * 1.5 + 1.5)
    return w, h


def _cate_scatter(ax, art, res_cols: list):
    if not res_cols:
        ax.scatter(art.T, art.cate, alpha=0.4, s=15, color="steelblue")
    elif len(res_cols) <= LEGEND_MAX:
        for rc in res_cols:
            mask = art.X[:, art.feature_cols.index(rc)] == 1
            ax.scatter(art.T[mask], art.cate[mask],
                       alpha=0.4, s=15, label=rc.replace("res_", ""))
        ax.legend(title="resource", fontsize=7, loc="best")
    else:
        res_index = np.full(len(art.T), -1, dtype=int)
        for k, rc in enumerate(res_cols):
            mask = art.X[:, art.feature_cols.index(rc)] == 1
            res_index[mask] = k
        cmap = plt.get_cmap("tab20", len(res_cols))
        sc = ax.scatter(art.T, art.cate, c=res_index, cmap=cmap,
                        vmin=0, vmax=len(res_cols) - 1, alpha=0.4, s=15)
        cb = plt.colorbar(sc, ax=ax, pad=0.01)
        cb.set_ticks(range(len(res_cols)))
        cb.set_ticklabels([rc.replace("res_", "") for rc in res_cols], fontsize=6)
        cb.set_label("resource", fontsize=7)
    ax.axhline(0, color="black", lw=0.8, ls="--")
    ax.set_xlabel("Utilisation (T)")
    ax.set_ylabel("CATE θ(X)")
    ax.set_title("CATE vs utilisation")


def _cate_statistics(artifacts) -> pd.DataFrame:
    rows = []
    for art in artifacts:
        c = art.cate
        rows.append({
            "choice_set":   art.choice_set,
            "transition":   art.transition,
            "n":            art.n,
            "ate":          round(float(c.mean()), 6),
            "cate_std":     round(float(c.std()),  6),
            "cate_min":     round(float(c.min()),  6),
            "cate_p25":     round(float(np.percentile(c, 25)), 6),
            "cate_p50":     round(float(np.percentile(c, 50)), 6),
            "cate_p75":     round(float(np.percentile(c, 75)), 6),
            "cate_max":     round(float(c.max()),  6),
            "pct_positive": round(float((c > 0).mean() * 100), 1),
            "pct_negative": round(float((c < 0).mean() * 100), 1),
            "top_feature":  art.feat_imp.idxmax() if art.feat_imp.sum() > 0 else "—",
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Excel formatting
# ─────────────────────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="2C3E50")   # dark blue header
_SIG_FILL  = PatternFill("solid", fgColor="D5F5E3")   # green  — significant / good
_WARN_FILL = PatternFill("solid", fgColor="FEF9E7")   # yellow — caution (△)
_BAD_FILL  = PatternFill("solid", fgColor="FDEBD0")   # orange — problem (✗)
_ALT_FILL  = PatternFill("solid", fgColor="F8F9FA")   # light gray — alternating
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
_BODY_FONT = Font(size=9)


def _format_sheet(
    ws,
    df: pd.DataFrame,
    sig_col:  str | None = None,
    flag_col: str | None = None,
) -> None:
    """
    Apply consistent formatting to an openpyxl worksheet.

    sig_col  : column name whose value True → green row highlight
    flag_col : column whose string value controls colour:
               contains "✓" → green, "△" → yellow, "✗" → orange
    """
    cols      = list(df.columns)
    sig_cidx  = (cols.index(sig_col)  + 1) if sig_col  and sig_col  in cols else None
    flag_cidx = (cols.index(flag_col) + 1) if flag_col and flag_col in cols else None

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx == 1:
            for cell in row:
                cell.fill      = _HDR_FILL
                cell.font      = _HDR_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            continue

        fill = None
        if sig_cidx:
            if row[sig_cidx - 1].value is True:
                fill = _SIG_FILL
        if fill is None and flag_cidx:
            flag_val = str(row[flag_cidx - 1].value or "")
            if "✓" in flag_val:
                fill = _SIG_FILL
            elif "△" in flag_val:
                fill = _WARN_FILL
            elif "✗" in flag_val:
                fill = _BAD_FILL
        if fill is None and row_idx % 2 == 0:
            fill = _ALT_FILL

        for cell in row:
            if fill:
                cell.fill = fill
            cell.font      = _BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

    ws.freeze_panes   = "A2"
    ws.auto_filter.ref = ws.dimensions


def _assign_cs_index(*dfs: pd.DataFrame) -> dict[str, int]:
    """Return an ordered mapping  choice_set_str → integer index.

    Order is determined by first appearance across the supplied DataFrames
    (typically dml_results, summary, backdoor).
    """
    seen: dict[str, int] = {}
    counter = 0
    for df in dfs:
        if df.empty or "choice_set" not in df.columns:
            continue
        for cs in df["choice_set"].astype(str):
            if cs not in seen:
                seen[cs] = counter
                counter += 1
    return seen


def _add_cs_idx(df: pd.DataFrame, cs_map: dict[str, int]) -> pd.DataFrame:
    """Prepend a cs_idx column (integer) to df."""
    out = df.copy()
    if "choice_set" in out.columns:
        out.insert(0, "cs_idx", out["choice_set"].astype(str).map(cs_map))
    return out


def _build_excel(
    dml_results: pd.DataFrame,
    summary:     pd.DataFrame,
    artifacts:   list,
    backdoor:    pd.DataFrame,
    path:        Path,
) -> None:
    """
    Write the findings workbook:

      CS_index      — mapping of cs_idx → choice_set string
      2_double_ml   — Double ML results with cs_idx; significant rows green
      3_forest_ate  — Causal Forest ATE with cs_idx; significant rows green
      3b_cate_stats — CATE distribution statistics with cs_idx
      BD_0, BD_1, … — one sheet per choice set for the backdoor checks,
                       ✓/△/✗ colour-coded
    """
    cate_stats = _cate_statistics(artifacts)

    # Build a consistent integer index across all result tables
    cs_map = _assign_cs_index(dml_results, summary, backdoor)

    # Add cs_idx column to the summary sheets
    dml_s  = _add_cs_idx(_stringify_lists(dml_results), cs_map)
    sum_s  = _add_cs_idx(_stringify_lists(summary),     cs_map)
    stat_s = _add_cs_idx(_stringify_lists(cate_stats),  cs_map)
    bd_s   = _add_cs_idx(_stringify_lists(backdoor),    cs_map)

    # Build the index lookup table
    index_df = pd.DataFrame(
        [{"cs_idx": idx, "choice_set": cs}
         for cs, idx in sorted(cs_map.items(), key=lambda kv: kv[1])]
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ── CS index sheet ────────────────────────────────────────────────────
        index_df.to_excel(writer, sheet_name="CS_index", index=False)
        _format_sheet(writer.sheets["CS_index"], index_df)

        # ── Summary sheets ────────────────────────────────────────────────────
        for sheet_name, df, sig_col, flag_col in [
            ("2_double_ml",   dml_s,  "reject_H0", None),
            ("3_forest_ate",  sum_s,  "reject_H0", None),
            ("3b_cate_stats", stat_s, None,        None),
        ]:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer.sheets[sheet_name], df,
                          sig_col=sig_col, flag_col=flag_col)

        # ── One backdoor sheet per choice set ─────────────────────────────────
        for cs_str, idx in sorted(cs_map.items(), key=lambda kv: kv[1]):
            subset = bd_s[bd_s["choice_set"] == cs_str]
            if subset.empty:
                continue
            sheet_name = f"BD_{idx}"          # e.g. BD_0, BD_1, …
            subset.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer.sheets[sheet_name], subset, flag_col="flag")


# ─────────────────────────────────────────────────────────────────────────────
# CATE tree PDFs — all transitions, split into significant / not-significant
# ─────────────────────────────────────────────────────────────────────────────

def _save_all_tree_pdfs(
    artifacts:      list,
    summary:        pd.DataFrame,
    sig_dir:        Path,
    not_sig_dir:    Path,
) -> dict[str, list[Path]]:
    """
    Save one PDF per CATE summary tree for every artifact, routing each file
    into sig_dir (reject_H0 == True) or not_sig_dir (reject_H0 != True).

    Returns {"significant": [...], "not_significant": [...]}.
    """
    sig_keys = set()
    if not summary.empty and "reject_H0" in summary.columns:
        sig      = summary[summary["reject_H0"] == True]
        sig_keys = set(zip(sig["choice_set"].astype(str), sig["transition"].astype(str)))

    result: dict[str, list[Path]] = {"significant": [], "not_significant": []}

    for art in artifacts:
        is_sig = (str(art.choice_set), str(art.transition)) in sig_keys

        safe = (
            str(art.choice_set)
            .replace("/", "-").replace(" ", "_")
            .replace("'", "").replace(",", "")
            .replace("(", "").replace(")", "")
        )
        dest  = sig_dir if is_sig else not_sig_dir
        fname = dest / f"cate_tree_{safe}__t{art.transition}.pdf"

        row   = summary[
            (summary["choice_set"].astype(str) == str(art.choice_set)) &
            (summary["transition"].astype(str) == str(art.transition))
        ] if not summary.empty else pd.DataFrame()
        ate  = float(row["ate"].iloc[0])   if not row.empty else float(art.cate.mean())
        ci   = str(row["ci_95"].iloc[0])   if not row.empty else "—"
        pval = str(row["p_value"].iloc[0]) if not row.empty else "—"

        fig_w, fig_h = _tree_figsize(art.tree)
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))
        plot_tree(
            art.tree,
            feature_names=art.feature_cols,
            filled=True, rounded=True, impurity=False, precision=4,
            ax=ax, fontsize=8,
        )
        sig_label = "SIGNIFICANT (p<0.05)" if is_sig else "not significant (p≥0.05)"
        ax.set_title(
            f"CATE summary tree  |  {sig_label}  |  choice_set={art.choice_set}  |  "
            f"transition='{art.transition}'  |  n={art.n}\n"
            f"ATE={ate:+.4f}  95% CI={ci}  p={pval}  "
            f"|  top feature: {art.feat_imp.idxmax()}\n"
            f"value > 0 → higher util raises P(chose '{art.transition}');  "
            f"value < 0 → lowers it;  value ≈ 0 → no effect",
            fontsize=9,
        )
        fig.tight_layout(pad=0.5)
        fig.savefig(str(fname), format="pdf", bbox_inches="tight")
        plt.close(fig)

        key = "significant" if is_sig else "not_significant"
        result[key].append(fname)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# PDF narrative report
# ─────────────────────────────────────────────────────────────────────────────

def _build_pdf(decisions, lr_results, dml_results, summary, artifacts,
               dr, backdoor, file_stem, path: Path):

    with PdfPages(path) as pdf:

        # ── Title page ────────────────────────────────────────────────────────
        fig = plt.figure(figsize=(11, 8.5))
        fig.patch.set_facecolor("#2c3e50")
        ax = fig.add_axes([0.1, 0.15, 0.8, 0.7])
        ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis("off")
        ax.text(0.5, 0.88, "Causal Analysis Report",
                ha="center", fontsize=26, fontweight="bold", color="white")
        ax.text(0.5, 0.76, file_stem,
                ha="center", fontsize=15, color="#bdc3c7")
        n_sig = int((summary["reject_H0"] == True).sum()) if not summary.empty else 0
        meta = [
            f"Decision points:      {len(decisions)}",
            f"Choice sets:          {decisions['choice_set'].nunique()}",
            f"Transitions fitted:   {len(artifacts)}",
            f"Significant (p<0.05): {n_sig}",
            f"Generated:            {datetime.now().strftime('%Y-%m-%d  %H:%M')}",
        ]
        for k, line in enumerate(meta):
            ax.text(0.5, 0.58 - k * 0.08, line,
                    ha="center", fontsize=12, color="#ecf0f1")
        ax.text(0.5, 0.06,
                "causal_analysis pipeline  ·  see ASSUMPTIONS.md for methodology",
                ha="center", fontsize=9, color="#7f8c8d")
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Step 1: Logistic regression ───────────────────────────────────────
        h = max(3.5, 0.45 * max(len(lr_results), 1) + 2)
        fig, ax = plt.subplots(figsize=(15, h))
        _render_table(
            ax, lr_results,
            "Step 1 — Association test  (logistic regression + permutation)\n"
            "H₀: choice ⊥ util | choice_set.   reject_H0=True → util predicts routing."
        )
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Step 2: Double ML ─────────────────────────────────────────────────
        h = max(3.5, 0.45 * max(len(dml_results), 1) + 2)
        fig, ax = plt.subplots(figsize=(15, h))
        _render_table(
            ax, dml_results,
            "Step 2 — Double ML  (partially linear regression, HC3 robust SE)\n"
            "θ = linear effect of 1-unit ↑ util on P(chose k), after partialling out X.\n"
            "Equals ATE when the effect is homogeneous. Otherwise: variance-weighted average of CATEs."
        )
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Step 3: Causal Forest — ATE table + bar chart ────────────────────
        summary_plot = summary[summary["ci_95"] != "—"].dropna(subset=["ate"]).copy()

        fig, axes = plt.subplots(
            1, 2, figsize=(16, max(4, 0.55 * max(len(summary_plot), 1) + 2.5)),
            gridspec_kw={"width_ratios": [1.4, 1]},
        )
        _render_table(
            axes[0], summary,
            "Step 3 — Causal Forest DML  (ATE per transition)\n"
            "ATE = mean of per-sample CATEs.  Green rows = H0 rejected (p<0.05).\n"
            "Less powerful for significance than Double ML, but basis for heterogeneity analysis."
        )
        if not summary_plot.empty:
            t_labels = summary_plot["transition"].tolist()
            ates_    = summary_plot["ate"].tolist()
            ci_los, ci_his = [], []
            for ci_str in summary_plot["ci_95"]:
                lo, hi = ci_str.strip("[]").split(",")
                ci_los.append(float(lo)); ci_his.append(float(hi))
            err_lo = [max(0.0, a - lo) for a, lo in zip(ates_, ci_los)]
            err_hi = [max(0.0, hi - a) for a, hi in zip(ates_, ci_his)]
            # mark significant bars darker
            sig_set = set(
                summary_plot.loc[summary_plot["reject_H0"] == True, "transition"].astype(str)
            )
            colors = []
            for t, a in zip(t_labels, ates_):
                if str(t) in sig_set:
                    colors.append("steelblue" if a >= 0 else "tomato")
                else:
                    colors.append("#aacfe4" if a >= 0 else "#f4b8b8")
            axes[1].barh(t_labels, ates_, xerr=[err_lo, err_hi],
                         color=colors, edgecolor="white", height=0.5,
                         error_kw=dict(ecolor="black", capsize=4))
            axes[1].axvline(0, color="black", lw=1)
            axes[1].set_xlabel("ATE  (util: 0 → 1)")
            axes[1].set_title(
                "ATE bar chart\nSaturated = p<0.05;  Faded = p≥0.05",
                fontsize=9,
            )
        else:
            axes[1].axis("off")
            axes[1].text(0.5, 0.5, "No valid ATEs to plot",
                         ha="center", va="center", transform=axes[1].transAxes, color="gray")
        fig.tight_layout()
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Per-transition: CATE plots + tree (all fitted transitions) ────────
        for art in artifacts:
            res_cols = [c for c in art.feature_cols if c.startswith("res_")]

            fig, axes = plt.subplots(1, 3, figsize=(16, 4))
            fig.suptitle(
                f"CATE — choice_set={art.choice_set}  |  transition='{art.transition}'  |  n={art.n}  |  "
                f"ATE={art.cate.mean():.4f}  std={art.cate.std():.4f}",
                fontsize=10, fontweight="bold",
            )
            _cate_scatter(axes[0], art, res_cols)

            axes[1].hist(art.cate, bins=40, color="steelblue", edgecolor="white")
            axes[1].axvline(art.cate.mean(), color="red", lw=2,
                            label=f"ATE={art.cate.mean():.4f}")
            axes[1].axvline(0, color="black", lw=0.8, ls="--")
            axes[1].set_xlabel("CATE θ(X)"); axes[1].set_ylabel("Count")
            axes[1].set_title("CATE distribution\nWide = heterogeneous;  Narrow = homogeneous")
            axes[1].legend(fontsize=8)

            res_labels_, means_, cis_ = [], [], []
            for rc in res_cols:
                idx_ = art.feature_cols.index(rc)
                mask = art.X[:, idx_] == 1
                if mask.sum() > 5:
                    res_labels_.append(rc.replace("res_", ""))
                    means_.append(art.cate[mask].mean())
                    cis_.append(art.cate[mask].std() / np.sqrt(mask.sum()))
            if means_:
                colors_r = ["steelblue" if m >= 0 else "tomato" for m in means_]
                axes[2].barh(res_labels_, means_, xerr=cis_, color=colors_r,
                             edgecolor="white", error_kw=dict(ecolor="black", capsize=3))
                axes[2].axvline(0, color="black", lw=0.8)
                axes[2].set_xlabel("Mean CATE")
                axes[2].set_title("Mean CATE by resource")
            else:
                axes[2].axis("off")
                axes[2].text(0.5, 0.5, "No resource columns",
                             ha="center", va="center", transform=axes[2].transAxes, color="gray")
            fig.tight_layout()
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

            fig_w, fig_h = _tree_figsize(art.tree)
            fig, ax = plt.subplots(figsize=(fig_w, fig_h))
            plot_tree(
                art.tree,
                feature_names=art.feature_cols,
                filled=True, rounded=True, impurity=False, precision=4,
                ax=ax, fontsize=8,
            )
            ax.set_title(
                f"CATE summary tree  |  choice_set={art.choice_set}  |  "
                f"transition='{art.transition}'  |  top feature: {art.feat_imp.idxmax()}\n"
                f"value > 0 → higher util raises P(chose '{art.transition}');  "
                f"value < 0 → lowers it;  value ≈ 0 → no effect",
                fontsize=9,
            )
            fig.tight_layout(pad=0.5)
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Step 4: Dose-response ─────────────────────────────────────────────
        if not dr.empty and artifacts:
            n_arts = len(artifacts)
            fig, axes = plt.subplots(1, n_arts, figsize=(6 * n_arts, 4), sharey=False)
            if n_arts == 1:
                axes = [axes]
            fig.suptitle(
                "Step 4 — Dose-response curves  (CADR)\n"
                "E[ΔP(chose k)] when util is set to t vs. util=0",
                fontsize=10, fontweight="bold",
            )
            for ax, art in zip(axes, artifacts):
                grp = dr[(dr["transition"] == art.transition) & (dr["subgroup"] == "all")]
                if grp.empty:
                    ax.text(0.5, 0.5, "No data", ha="center", va="center",
                            transform=ax.transAxes, color="gray")
                    continue
                ax.plot(grp["t"], grp["mean_effect"], color="steelblue")
                ax.fill_between(grp["t"], grp["ci_lb"], grp["ci_ub"],
                                alpha=0.2, color="steelblue", label="95% CI")
                ax.axhline(0, color="black", lw=0.8, ls="--")
                ax.set_xlabel("util (t)"); ax.set_ylabel("Δ P(chose k) vs. util=0")
                ax.set_title(f"transition='{art.transition}'")
                ax.legend(fontsize=8)
            fig.tight_layout()
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Step 5: Backdoor check ────────────────────────────────────────────
        if not backdoor.empty:
            h = max(3.5, 0.38 * max(len(backdoor), 1) + 2)
            fig, ax = plt.subplots(figsize=(15, h))
            _render_table(
                ax, backdoor,
                "Step 5 — Backdoor adequacy\n"
                "nuisance_R²: does X predict util? (high is good)  |  "
                "placebo: spurious signal? (should be ≈0)  |  SMD: |value|<0.1 = balanced"
            )
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── CATE statistics summary ───────────────────────────────────────────
        cate_stats = _cate_statistics(artifacts)
        if not cate_stats.empty:
            h = max(3.5, 0.45 * max(len(cate_stats), 1) + 2)
            fig, ax = plt.subplots(figsize=(15, h))
            _render_table(
                ax, cate_stats,
                "CATE statistics summary\n"
                "cate_std = heterogeneity;  pct_positive/negative = share of cases where util helps/hurts;  "
                "top_feature = main moderator"
            )
            pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        # ── Assumptions ───────────────────────────────────────────────────────
        assumptions = [
            ("Backdoor criterion  (all methods)",
             "X = {hour, dayofweek, n_alternatives, res_*} must block all confounding paths T ← C → Y.\n"
             "Any variable that affects both util and routing must be in X. Unobserved confounders\n"
             "cannot be detected and will bias all estimates."),
            ("SUTVA  (all methods)",
             "One case's routing does not affect another's treatment or outcome.\n"
             "Violated if resource occupation by one case changes the util seen by a concurrent case."),
            ("Positivity  (all methods)",
             "Every covariate combination X must have positive probability for all util levels T.\n"
             "Diagnosed by the covariate balance (SMD) in the backdoor check."),
            ("Linearity  (Double ML only)",
             "θ is the best-fit linear slope. If the effect varies with X, θ is a variance-weighted\n"
             "average of CATEs — valid as a summary, but not the simple mean.\n"
             "Double ML is more powerful than the forest for testing 'is there any effect?'"),
            ("Honest splitting  (Causal Forest only)",
             "Tree structure and leaf estimates use separate subsamples. Yields valid CIs but\n"
             "halves effective sample size → forest CIs are wider than Double ML's.\n"
             "A significant Double ML θ with an insignificant forest ATE is not a contradiction."),
            ("What the analysis cannot establish",
             "Causal direction (assumes util → choice, not the reverse).\n"
             "External validity beyond the observed covariate distribution.\n"
             "Dynamic/lagged effects across cases over time."),
        ]
        fig, ax = plt.subplots(figsize=(13, 10))
        ax.axis("off")
        ax.text(0.0, 0.99, "Key Assumptions & Interpretation Notes",
                fontsize=13, fontweight="bold", transform=ax.transAxes, va="top")
        ax.plot([0, 1], [0.965, 0.965], color="#2c3e50", linewidth=1.5,
                transform=ax.transAxes, clip_on=False)
        y = 0.93
        for name, body in assumptions:
            ax.text(0.0, y, f"▸  {name}", fontsize=9, fontweight="bold",
                    transform=ax.transAxes, va="top", color="#2c3e50")
            y -= 0.04
            for line in body.split("\n"):
                ax.text(0.03, y, line, fontsize=8.5, color="#444444",
                        transform=ax.transAxes, va="top")
                y -= 0.035
            y -= 0.015
        ax.text(0.0, 0.02, "Full documentation: ASSUMPTIONS.md",
                fontsize=8, color="gray", transform=ax.transAxes)
        pdf.savefig(fig, bbox_inches="tight"); plt.close(fig)

        d = pdf.infodict()
        d["Title"]   = f"Causal Analysis Report — {file_stem}"
        d["Subject"] = "Resource utilisation → routing decision (causal inference)"
        d["Creator"] = "reporting.py / causal_analysis pipeline"


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def generate_report(
    decisions:   pd.DataFrame,
    lr_results:  pd.DataFrame,
    dml_results: pd.DataFrame,
    summary:     pd.DataFrame,
    artifacts:   list,
    dr:          pd.DataFrame,
    backdoor:    pd.DataFrame,
    output_dir:  str | os.PathLike,
    file_stem:   str = "report",
) -> dict:
    """
    Generate findings Excel, CATE tree PDFs, and PDF narrative report.

    Excel  — {file_stem}.xlsx
        Sheets: 2_double_ml, 3_forest_ate, 3b_cate_stats, 5_backdoor
        Rows highlighted: green = significant / good, yellow = caution, orange = problem

    Tree PDFs — cate_trees/significant/ and cate_trees/not-significant/
        One PDF per (choice_set, transition) for every fitted artifact.
        Significant (p < 0.05) trees go into the significant/ subfolder;
        all others go into not-significant/.

    PDF report — {file_stem}.pdf
        Full narrative report: all tables, CATE plots, dose-response curves,
        assumptions page.

    Returns
    -------
    dict with keys:
        "excel"                      — path to the Excel workbook
        "pdf"                        — path to the narrative PDF
        "trees_dir"                  — path to the cate_trees/ root folder
        "tree_pdfs_significant"      — list of paths in cate_trees/significant/
        "tree_pdfs_not_significant"  — list of paths in cate_trees/not-significant/
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    trees_dir     = output_dir / "cate_trees"
    sig_dir       = trees_dir / "significant"
    not_sig_dir   = trees_dir / "not-significant"
    for d in (trees_dir, sig_dir, not_sig_dir):
        d.mkdir(exist_ok=True)

    excel_path = output_dir / f"{file_stem}.xlsx"
    pdf_path   = output_dir / f"{file_stem}.pdf"

    _build_excel(dml_results, summary, artifacts, backdoor, excel_path)

    tree_pdfs = _save_all_tree_pdfs(artifacts, summary, sig_dir, not_sig_dir)

    _build_pdf(decisions, lr_results, dml_results, summary, artifacts,
               dr, backdoor, file_stem, pdf_path)

    return {
        "excel":                str(excel_path),
        "pdf":                  str(pdf_path),
        "trees_dir":            str(trees_dir),
        "tree_pdfs_significant":     [str(p) for p in tree_pdfs["significant"]],
        "tree_pdfs_not_significant": [str(p) for p in tree_pdfs["not_significant"]],
    }
